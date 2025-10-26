
#---Función 1: Control de inventario---#
import os
import tkinter as tk
from tkinter import messagebox, ttk, simpledialog
import openpyxl
from openpyxl import Workbook
import subprocess

# Definir la ruta del archivo de inventario
control_de_inventario = "inventario.xlsx"

# Función para crear el libro de inventario si no existe
def crear_libro(path):
    """Crea un nuevo libro con encabezados si no existe."""
    if os.path.exists(path):
        return
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Inventario"
    hoja['A1'] = 'Codigo'
    hoja['B1'] = 'Nombre'
    hoja['C1'] = 'Existencia'
    hoja['D1'] = 'Proveedor'
    hoja['E1'] = 'Precio'
    wb.save(path)

# Función para abrir el libro de inventario
def abrir_libro(path):
    """Abre el libro, crea si no existe."""
    if not os.path.exists(path):
        crear_libro(path)
    return openpyxl.load_workbook(path)

def listar_productos():
    libro = abrir_libro(control_de_inventario)
    hoja = libro.active

    productos = []

    # si solo está la fila de encabezado, no hay productos
    if hoja.max_row < 2:
        return productos

    # Recorrer filas desde la 2 (después del encabezado)
    for fila in range(2, hoja.max_row + 1):
        codigo = hoja[f"A{fila}"].value
        nombre = hoja[f"B{fila}"].value
        existencia = hoja[f"C{fila}"].value
        proveedor = hoja[f"D{fila}"].value
        precio = hoja[f"E{fila}"].value
        linea = f"{codigo} | {nombre} | Exist: {existencia} | Prov: {proveedor} | Precio: {precio}"
        productos.append(linea)
    return productos

# Función para mostrar el listado de productos en una nueva ventana
def mostrar_listado():
    productos = listar_productos()
    ventana_listado = tk.Toplevel(ventana)
    ventana_listado.title("Listado de Productos")
    ventana_listado.geometry("600x400")

    # Si no hay productos, mostrar mensaje
    if not productos:
        tk.Label(ventana_listado, text="No hay productos en el inventario.").pack(padx=10, pady=10)
        return

    frame = tk.Frame(ventana_listado)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")

    listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, font=("Consolas", 10))
    for p in productos:
        listbox.insert(tk.END, p)
    listbox.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=listbox.yview)

# Función para crear un nuevo producto
def crear_producto():
    try:
        libro = abrir_libro(control_de_inventario)
        hoja = libro.active
    except Exception as e:
        messagebox.showerror("Error", f"No se puede abrir el archivo: {e}")
        return

    codigo = entrada_codigo.get().strip()
    nombre = entrada_nombre.get().strip()
    existencia = entrada_existencia.get().strip()
    proveedor = entrada_proveedor.get().strip()
    precio = entrada_precio.get().strip()

    if not (codigo and nombre and existencia and proveedor and precio):
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        return

    # verificar existencia y precio
    try:
        existencia_val = int(existencia)
    except ValueError:
        try:
            existencia_val = float(existencia)
        except ValueError:
            messagebox.showwarning("Advertencia", "Existencia debe ser un número")
            return

    try:
        precio_val = float(precio)
    except ValueError:
        messagebox.showwarning("Advertencia", "Precio debe ser un número (usar punto como separador decimal)")
        return

    # Agregar nueva fila
    nueva_fila = hoja.max_row + 1
    hoja[f"A{nueva_fila}"] = codigo
    hoja[f"B{nueva_fila}"] = nombre
    hoja[f"C{nueva_fila}"] = existencia_val
    hoja[f"D{nueva_fila}"] = proveedor
    hoja[f"E{nueva_fila}"] = precio_val

    # Guardar cambios
    try:
        libro.save(control_de_inventario)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
        return

# Confirmación y limpieza de campos
    messagebox.showinfo("Éxito", "Producto creado exitosamente")
    entrada_codigo.delete(0, tk.END)
    entrada_nombre.delete(0, tk.END)
    entrada_existencia.delete(0, tk.END)
    entrada_proveedor.delete(0, tk.END)
    entrada_precio.delete(0, tk.END)

# Función para editar productos existentes
def editar_producto(event=None):
    seleccion = simpledialog.askstring("Editar Producto", "Ingrese el código del producto a editar:")
    if not seleccion:
        return
    indice = None
    libro = abrir_libro(control_de_inventario)
    hoja = libro.active
    # Buscar el índice del producto
    for fila in range(2, hoja.max_row + 1):
        if str(hoja[f"A{fila}"].value) == seleccion:
            indice = fila
            break
    # Si no se encuentra el producto
    if indice is None:
        messagebox.showerror("Error", "Producto no encontrado")
        return
    # Crear ventana de edición
    ventana_edicion = tk.Toplevel(ventana)
    ventana_edicion.title("Editar Producto")
    ventana_edicion.geometry("300x250")
    # Campos de edición
    tk.Label(ventana_edicion, text="Código:").grid(row=0, column=0, sticky="w", pady=2)
    entrada_codigo_edit = tk.Entry(ventana_edicion) 
    entrada_codigo_edit.grid(row=0, column=1, sticky="ew", pady=2)
    entrada_codigo_edit.insert(0, hoja[f"A{indice}"].value)
    tk.Label(ventana_edicion, text="Nombre:").grid(row=1, column=0, sticky="w", pady=2)
    entrada_nombre_edit = tk.Entry(ventana_edicion)
    entrada_nombre_edit.grid(row=1, column=1, sticky="ew", pady=2)
    entrada_nombre_edit.insert(0, hoja[f"B{indice}"].value)
    tk.Label(ventana_edicion, text="Existencia:").grid(row=2, column=0, sticky="w", pady=2)
    entrada_existencia_edit = tk.Entry(ventana_edicion)
    entrada_existencia_edit.grid(row=2, column=1, sticky="ew", pady=2)
    entrada_existencia_edit.insert(0, hoja[f"C{indice}"].value)
    tk.Label(ventana_edicion, text="Proveedor:").grid(row=3, column=0, sticky="w", pady=2)
    entrada_proveedor_edit = tk.Entry(ventana_edicion)
    entrada_proveedor_edit.grid(row=3, column=1, sticky="ew", pady=2)
    entrada_proveedor_edit.insert(0, hoja[f"D{indice}"].value)
    tk.Label(ventana_edicion, text="Precio:").grid(row=4, column=0, sticky="w", pady=2)
    entrada_precio_edit = tk.Entry(ventana_edicion)
    entrada_precio_edit.grid(row=4, column=1, sticky="ew", pady=2)
    entrada_precio_edit.insert(0, hoja[f"E{indice}"].value)
    # Ajustar columnas
    ventana_edicion.columnconfigure(1, weight=1)
    # Función para guardar cambios
    def guardar_cambios():
        # Obtener nuevos valores
        nuevo_codigo = entrada_codigo_edit.get().strip()
        nuevo_nombre = entrada_nombre_edit.get().strip()
        nueva_existencia = entrada_existencia_edit.get().strip()
        nuevo_proveedor = entrada_proveedor_edit.get().strip()
        nuevo_precio = entrada_precio_edit.get().strip()
        if not (nuevo_codigo and nuevo_nombre and nueva_existencia and nuevo_proveedor and nuevo_precio):
            messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
            return
        # verificar existencia
        try:
            nueva_existencia_val = int(nueva_existencia)
        except ValueError:
            try:
                nueva_existencia_val = float(nueva_existencia)
            except ValueError:
                messagebox.showwarning("Advertencia", "Existencia debe ser un número")
                return
        # verificar precio
        try:
            nuevo_precio_val = float(nuevo_precio)
        except ValueError:
            messagebox.showwarning("Advertencia", "Precio debe ser un número (usar punto como separador decimal)")
            return
        # Actualizar valores en la hoja
        hoja[f"A{indice}"] = nuevo_codigo
        hoja[f"B{indice}"] = nuevo_nombre
        hoja[f"C{indice}"] = nueva_existencia_val
        hoja[f"D{indice}"] = nuevo_proveedor
        hoja[f"E{indice}"] = nuevo_precio_val
        # Guardar cambios
        try:
            libro.save(control_de_inventario)
            messagebox.showinfo("Éxito", "Producto actualizado exitosamente")
            ventana_edicion.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
            # No cerrar la ventana si hay error
    btn_guardar = tk.Button(ventana_edicion, text="Guardar Cambios", command=guardar_cambios)
    btn_guardar.grid(row=5, column=0, columnspan=2, pady=10)

#funcion para eliminar productos
def eliminar_producto(event=None):
    seleccion = simpledialog.askstring("Eliminar Producto", "Ingrese el código del producto a eliminar:")
    if not seleccion:
        return
    # Buscar el índice del producto
    indice = None
    libro = abrir_libro(control_de_inventario)
    hoja = libro.active
    for fila in range(2, hoja.max_row + 1):
        if str(hoja[f"A{fila}"].value) == seleccion:
            indice = fila
            break
    if indice is None:
        messagebox.showerror("Error", "Producto no encontrado")
        return
    confirmacion = messagebox.askyesno("Confirmar", f"¿Está seguro de eliminar el producto con código {seleccion}?")
    if not confirmacion:
        return
    # Eliminar la fila
    hoja.delete_rows(indice)
    try:
        libro.save(control_de_inventario)
        messagebox.showinfo("Éxito", "Producto eliminado exitosamente")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
#Interfaz principal (único menú)
crear_libro(control_de_inventario)  # corroborar existencia del archivo

def boton_regresar():
    ventana.destroy()
    subprocess.run(['python', 'PROYECTO_FINAL/Algoritmos/main_algoritmos.py'])

# Configuración de la ventana principal
ventana = tk.Tk() 
ventana.state('zoomed')
ventana.title("Control de Inventario")
ventana.geometry("640x380")
ventana.resizable(True, True)

# Estilo
style = ttk.Style()

# Ajustes simples de estilo
try:
    style.configure("TLabelFrame", padding=8)
    style.configure("TButton", padding=6)
except Exception:
    pass

# Contenedor principal
main_frame = ttk.Frame(ventana, padding=12)
main_frame.grid(sticky="nsew")
ventana.columnconfigure(0, weight=1)
ventana.rowconfigure(0, weight=1)
main_frame.columnconfigure(0, weight=1)
main_frame.columnconfigure(1, weight=0)

# Título
titulo = ttk.Label(main_frame, text="Control de Inventario", font=("Helvetica", 16, "bold"))
titulo.grid(row=0, column=0, columnspan=2, pady=(0, 12), sticky="w")

# Formulario agrupado
form_frame = ttk.LabelFrame(main_frame, text="Formulario")
form_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 12), pady=4)
form_frame.columnconfigure(1, weight=1)

ttk.Label(form_frame, text="Código:").grid(row=0, column=0, sticky="w", pady=6, padx=(0,6))
entrada_codigo = ttk.Entry(form_frame)
entrada_codigo.grid(row=0, column=1, sticky="ew", pady=6)

ttk.Label(form_frame, text="Nombre:").grid(row=1, column=0, sticky="w", pady=6, padx=(0,6))
entrada_nombre = ttk.Entry(form_frame)
entrada_nombre.grid(row=1, column=1, sticky="ew", pady=6)

ttk.Label(form_frame, text="Existencia:").grid(row=2, column=0, sticky="w", pady=6, padx=(0,6))
entrada_existencia = ttk.Entry(form_frame)
entrada_existencia.grid(row=2, column=1, sticky="ew", pady=6)

ttk.Label(form_frame, text="Proveedor:").grid(row=3, column=0, sticky="w", pady=6, padx=(0,6))
entrada_proveedor = ttk.Entry(form_frame)
entrada_proveedor.grid(row=3, column=1, sticky="ew", pady=6)

ttk.Label(form_frame, text="Precio:").grid(row=4, column=0, sticky="w", pady=6, padx=(0,6))
entrada_precio = ttk.Entry(form_frame)
entrada_precio.grid(row=4, column=1, sticky="ew", pady=6)

# Panel de acciones
acciones_frame = ttk.LabelFrame(main_frame, text="Acciones")
acciones_frame.grid(row=1, column=1, sticky="n", pady=4)
for i in range(2): acciones_frame.columnconfigure(i, weight=1)

btn_crear = ttk.Button(acciones_frame, text="Crear producto", width=18, command=crear_producto)
btn_crear.grid(row=0, column=0, columnspan=2, pady=(4,6), padx=6, sticky="ew")

btn_listar = ttk.Button(acciones_frame, text="Listar productos", width=18, command=mostrar_listado)
btn_listar.grid(row=1, column=0, pady=4, padx=6, sticky="ew")

btn_editar = ttk.Button(acciones_frame, text="Editar producto", width=18, command=editar_producto)
btn_editar.grid(row=1, column=1, pady=4, padx=6, sticky="ew")

btn_eliminar = ttk.Button(acciones_frame, text="Eliminar producto", width=18, command=eliminar_producto)
btn_eliminar.grid(row=2, column=0, pady=4, padx=6, sticky="ew")

btn_recrear = ttk.Button(acciones_frame, text="Recrear encabezados", width=18, command=lambda: (crear_libro(control_de_inventario), messagebox.showinfo("Listo", "Archivo creado/revisado")))
btn_recrear.grid(row=2, column=1, pady=4, padx=6, sticky="ew")

btn_regresar = ttk.Button(main_frame, text="Regresar al menú principal", command=boton_regresar)
btn_regresar.grid(row=3, column=0, columnspan=2, pady=(12,0), sticky="ew")

# Información al pie
info = ttk.Label(main_frame, text="Los productos se guardan en 'inventario.xlsx'", foreground="gray")
info.grid(row=2, column=0, columnspan=2, pady=(12,0), sticky="w")

# Iniciar el bucle principal de la interfaz
ventana.mainloop()